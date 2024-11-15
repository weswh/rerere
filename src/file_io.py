import os
import cv2
import numpy as np
import json 
import cv2 as cv
from tqdm import tqdm
from pathlib import Path
from typing import Union, Tuple
from openpyxl import Workbook, load_workbook
import pyvista as pv
import trimesh
import open3d as o3d


def filter_and_save_left_model(ply_file_path, obj_file_path):
    # 使用 Open3D 读取 .ply 文件
    mesh = o3d.io.read_triangle_mesh(ply_file_path)
    # 修剪模型
    vertices = np.asarray(mesh.vertices)
    vertices_to_keep = vertices[:, 0] <= 0
    # print('vertices_to_keep',vertices_to_keep)
    triangles_to_remove = np.any(vertices_to_keep[mesh.triangles], axis=1)
    mesh.remove_triangles_by_mask(triangles_to_remove)
    mesh.remove_unreferenced_vertices()
    mesh.remove_degenerate_triangles()
    o3d.io.write_triangle_mesh(obj_file_path, mesh)

def filter_and_save_right_model(ply_file_path, obj_file_path):
    # 使用 Open3D 读取 .ply 文件
    mesh = o3d.io.read_triangle_mesh(ply_file_path)
    # 修剪模型
    vertices = np.asarray(mesh.vertices)
    vertices_to_keep = vertices[:, 0] >= 0
    # print('vertices_to_keep',vertices_to_keep)
    triangles_to_remove = np.any(vertices_to_keep[mesh.triangles], axis=1)
    mesh.remove_triangles_by_mask(triangles_to_remove)
    mesh.remove_unreferenced_vertices()
    mesh.remove_degenerate_triangles()
    # 获取修剪后的顶点
    vertices = np.asarray(mesh.vertices)
    # 镜像操作：对 X 轴的坐标取反
    vertices[:, 0] = -vertices[:, 0]
    # 更新 mesh 顶点
    mesh.vertices = o3d.utility.Vector3dVector(vertices)
    o3d.io.write_triangle_mesh(obj_file_path, mesh)

def save_data_to_excel(file_path, data):
    # 检查Excel文件是否存在
    if os.path.exists(file_path):
        # 如果文件存在，加载文件
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # 如果文件不存在，创建一个新的Workbook
        workbook = Workbook()
        sheet = workbook.active
        # 写入表头（可选）
        sheet.append(["Data 1", "Data 2", "Data 3", "Data 4"])

    # 将新数据追加到Excel表中
    sheet.append(data)
    
    # 保存文件
    workbook.save(file_path)
    print(f"Data saved to {file_path}")

# 从已有文件夹读入images
def read_images_from_folder(folder_path, suffix=".jpg"):
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Folder '{folder_path}' does not exist.")
    names = []
    images = []
    shapes = []
    
    image_files = [f for f in os.listdir(folder_path) if f.endswith(suffix)]
    image_files.sort()

    for name in image_files:
        img_path = os.path.join(folder_path, name)
        image = cv2.imread(img_path)
        if image is None:
            print(f"Warning: Failed to load image '{img_path}'. Skipping.")
            continue
        images.append(image)
        shapes.append(image.shape[:2])
        names.append(name)
        
    h, w, _ = images[0].shape
    # h, w= images[0].shape
    n_images = len(images)
    images_array = np.zeros((n_images, h, w, 3), dtype=np.uint8)
    for i, img in enumerate(images):
        if img.shape[:2] != (h, w):
            raise ValueError(f"Image '{names[i]}' has different shape {img.shape[:2]} than expected {h, w}.")
        images_array[i] = img

    return np.asarray(names), images_array, np.asarray(shapes, dtype=np.int32)


# 保存图像
def save_image(path: Union[str, Path], image: np.ndarray) -> None:
    """ Save the image.

    Args:
        path: The image save path.
        image: The image data that needs to be saved.
    """
    if not isinstance(path, Path):
        path = Path(path)
    if not path.parent.is_dir():
        path.parent.mkdir(exist_ok=True, parents=True)
    cv.imwrite(str(path), image)
    # print("Save path: ", str(path))

# 保存图像
def save_images_to_dir(path: Union[str, Path], names: Union[list, np.ndarray], images: Union[list, np.ndarray]) -> None:
    """ Save all images to a folder.

    Args:
        path: The save folder path.
        names: The images name that needs to be saved.
        images: The images data that needs to be saved.

    Returns:

    """
    if not isinstance(path, Path):
        path = Path(path)
    for n, img in zip(names, images):
        img_pth = Path(path, n)
        save_image(img_pth, img)
    print("Save path: ", str(path))

# 存储为json文件
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(
            obj,
            (
                np.int_,
                np.intc,
                np.intp,
                np.int8,
                np.int16,
                np.int32,
                np.int64,
                np.uint8,
                np.uint16,
                np.uint32,
                np.uint64,
            ),
        ):
            return int(obj)
        elif isinstance(obj, (np.float_, np.float16, np.float32, np.float64)):
            return float(obj)
        elif isinstance(obj, (np.ndarray,)):
            return obj.tolist()
        return json.JSONEncoder.default(self, obj)
    
def save_json(path: Union[str, Path], data: dict) -> None:
    """Save the data in json format.

    Args:
        path (Union[str, Path]): The save path.
        data (dict): Need save data.
    """
    if not isinstance(path, Path):
        path = Path(path)
    if not path.parent.is_dir():
        path.parent.mkdir(exist_ok=True, parents=True)
    with open(str(path), "w") as f:
        json.dump(data, f, indent=4, cls=NumpyEncoder)
    print("The Json Path: ", str(path))
